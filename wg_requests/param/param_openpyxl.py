# -*- coding:utf8 -*-
import json
import openpyxl


class Param(object):
    def __init__(self, paramConf='{}'):
        self.paramConf = json.loads(paramConf)

    def paramRowsCount(self):
        pass

    def paramColsCount(self):
        pass

    def paramHeader(self):
        pass

    def paramAllline(self):
        pass

    def paramAlllineDict(self):
        pass


class XLSX(Param):
    '''
    xlsx基本格式

    第一行是参数的注释,就是每一行参数是什么

    第二行是参数名,参数名和对应模块的po页面的变量名一致

    第3~N行是参数

    最后一列是预期默认头Exp
    '''

    def __init__(self, paramConf):
        '''
        :param paramConf: xls 文件位置(绝对路径)
        '''
        self.paramConf = paramConf
        self.paramfile = self.paramConf['file']
        self.data = openpyxl.load_workbook(self.paramfile)
        # 与xlrd 模块的区别
        # wokrbook=xlrd.open_workbook(""DataSource\Economics.xls)
        self.getParamSheet(self.paramConf['sheet'])

    def getParamSheet(self, nsheets):
        '''
        设定参数所处的sheet
        :param nsheets: 参数在第几个sheet中
        :return:
        '''
        # self.paramsheet = self.data.get_sheet_names()[nsheets]
        sheet_names = self.data.get_sheet_names()
        self.paramsheet=self.data.get_sheet_by_name(sheet_names[nsheets])

    
    def getOneline(self, nRow):
        '''
        返回一行数据
        :param nRow: 行数
        :return: 一行数据 []
        '''
        max_column= self.paramsheet.max_column
        rowdata=[]
        for i in range(1,max_column+1):
            cellvalue=self.paramsheet.cell(row=nRow,column=i).value
            rowdata.append(cellvalue)
        return rowdata


    def getOneCol(self, nCol):
        '''
        返回一列
        :param nCol: 列数
        :return: 一列数据 []
        '''
        max_row = self.paramsheet.max_row
        columndata = []
        for i in range(1,max_row+1):
            cellvalue = self.paramsheet.cell(row=i, column=nCol).value
            columndata.append(cellvalue)
        return columndata

    def paramRowsCount(self):
        '''
        获取参数文件行数
        :return: 参数行数 int
        '''
        max_row=self.paramsheet.max_row
        return max_row

    def paramColsCount(self):
        '''
        获取参数文件列数(参数个数)
        :return: 参数文件列数(参数个数) int
        '''
        max_column=self.paramsheet.max_column
        return max_column

    def paramHeader(self):
        '''
        获取参数名称
        :return: 参数名称[]
        '''
        return self.getOneline(1)

    def paramAlllineDict(self):
        '''
        获取全部参数
        :return: {{}},其中dict的key值是header的值
        '''
        nCountRows = self.paramRowsCount()
        nCountCols = self.paramColsCount()
        ParamAllListDict = {}
        iRowStep = 2
        iColStep = 0
        
        ParamHeader = self.paramHeader()

        while iRowStep <= nCountRows:
            ParamOnelineDict = {}
            ParamOneLinelist = self.getOneline(iRowStep)
            print(ParamOneLinelist)

            while iColStep < nCountCols:
                ParamOnelineDict[ParamHeader[iColStep]] = ParamOneLinelist[iColStep]
                iColStep = iColStep + 1
            iColStep = 0
            ParamAllListDict[iRowStep - 2] = ParamOnelineDict
            iRowStep = iRowStep + 1
        print(ParamAllListDict)
        return ParamAllListDict

    def paramAllline(self):
        '''
        获取全部参数
        :return: 全部参数[[]]
        '''
        nCountRows = self.paramRowsCount()
        paramall = []
        iRowStep = 2
        while iRowStep <= nCountRows:
            paramall.append(self.getOneline(iRowStep))
            iRowStep = iRowStep + 1
        return paramall

    def __getParamCell(self, numberRow, numberCol):
        return self.paramsheet.cell_value(numberRow, numberCol)



class ParamFactory(object):
    def chooseParam(self, type, paramConf):
        map_ = {
            'xlsx': XLSX(paramConf)
        }
        return map_[type]